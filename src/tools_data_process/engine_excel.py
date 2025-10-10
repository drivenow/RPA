import pandas as pd
import sys
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


class ExcelEngine:
    def __init__(self, base_dir="X:\\RAG\\rpa_data\\"):
        self.base_dir = base_dir

    def dict_to_excel(self, data, file_name, sheet_name):
        # 第一步，将数据写入子excel文件
        df = pd.DataFrame(data)
        date = file_name.split('.')[0].split('_')[-1]
        df.insert(0, '重要性', "")
        df.insert(0, '更新日期', date)
        df = df.reset_index(drop=True)
        try:
            df.to_excel(file_name, sheet_name=sheet_name, index=False)
        except Exception as e:
            # 删除文件
            os.remove(file_name)
            df.to_excel(file_name, sheet_name=sheet_name, index=False)

    def get_excel_sheet_names(self, file_path: str) -> list:
        """
        获取Excel文件中所有工作表的名称

        参数:
        file_path: Excel文件路径

        返回:
        list: 工作表名称列表
        """
        # import fastexcel
        # target_sheets = fastexcel.read_excel(main_file_name).sheet_names
        try:
            # 加载工作簿 (只读模式可以提高性能)
            workbook = load_workbook(filename=file_path, read_only=True)

            # 方法1：直接获取 sheetnames 属性
            sheet_names = workbook.sheetnames

            # 方法2：从 worksheets 获取名称
            # sheet_names = [sheet.title for sheet in workbook.worksheets]

            # 关闭工作簿
            workbook.close()

            return sheet_names

        except Exception as e:
            print(f"读取Excel出错: {str(e)}")
            raise

    def get_summary_path(self, platform: str, date=None):
        if not os.path.exists(os.path.join(self.base_dir, platform)):
            print(f"WARNIGN: 创建新目录, {os.path.join(self.base_dir, platform)}")
            os.makedirs(os.path.join(self.base_dir, platform))
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
        else:
            date = pd.to_datetime(date).strftime('%Y-%m-%d')
        sub_file_path = os.path.join(self.base_dir, platform, f"{platform}_{date}.xlsx")
        main_file_path = os.path.join(self.base_dir, platform, f"{platform}_summary.xlsx")
        print("get_summary_path: ", sub_file_path, main_file_path)
        return sub_file_path, main_file_path

  

    def df_to_excel(self, df, file_name, sheet_name):
        date = file_name.split('.')[0].split('_')[-1]
        df.insert(0, '重要性', "")
        df.insert(0, '更新日期', date)
        df = df.reset_index(drop=True)
        try:
            df.to_excel(file_name, sheet_name=sheet_name, index=False)
        except Exception as e:
            # 删除文件
            os.remove(file_name)
            df.to_excel(file_name, sheet_name=sheet_name, index=False)

    def dict_df_to_excel(self, df_dict, file_name):
        # 异常若删除文件
        try:
            pd.read_excel(file_name, index=False)
        except Exception as e:
            try:
                os.remove(file_name)
            except:
                pass

        if not df_dict:
            return

        if not Path(file_name).exists():
            os.makedirs(os.path.dirname(file_name), exist_ok=True)
            empty_df = pd.DataFrame()
            with pd.ExcelWriter(file_name) as writer:
                # 将空的DataFrame写入文件
                empty_df.to_excel(writer, sheet_name=list(df_dict.keys())[0], index=False)

        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            date = file_name.split('.')[0].split('_')[-1]
            for sheet_name, df in df_dict.items():
                df.insert(0, '重要性', "")
                df.insert(0, '更新日期', date)
                df = df.reset_index(drop=True)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        self.format_excel(file_name)

    def merge_excel(self, sub_file_name, main_file_name, drop_duplicates_columns):
        """
        读取源Excel文件的数据并合并到目标Excel文件中，按sheet名进行覆盖

        参数:
        sub_file_name: 源Excel文件路径
        main_file_name: 目标Excel文件路径
        """
        if type(drop_duplicates_columns) == str:
            drop_duplicates_columns = [drop_duplicates_columns.strip()]
        drop_duplicates_columns = [i.strip() for i in drop_duplicates_columns]
        try:
            if not Path(sub_file_name).exists():
                print(f"源文件不存在: {sub_file_name}")
                return
            # 首先写入源文件中的sheet
            source_sheets = self.get_excel_sheet_names(sub_file_name)
            # 如果目标文件存在，读取现有的sheet
            if Path(main_file_name).exists():
                target_sheets = self.get_excel_sheet_names(main_file_name)
                # target_sheets
                existing_data = {}
                for sheet_name in target_sheets:
                    existing_data[sheet_name] = pd.read_excel(main_file_name, sheet_name=sheet_name)
            else:
                # 创建一个空的目标excel文件
                print(f"目标文件不存在，创建新文件: {main_file_name}")
                os.makedirs(os.path.dirname(main_file_name), exist_ok=True)
                # 创建一个空的DataFrame
                empty_df = pd.DataFrame()

                # 使用ExcelWriter创建一个Excel文件
                with pd.ExcelWriter(main_file_name) as writer:
                    # 将空的DataFrame写入文件
                    empty_df.to_excel(writer, sheet_name=source_sheets[0], index=False)
                target_sheets = []
                existing_data = {}

            # 读取源文件的数据并更新到目标数据中
            with pd.ExcelWriter(main_file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                for sheet_name in source_sheets:
                    df = pd.read_excel(sub_file_name, sheet_name=sheet_name)
                    # print(target_sheets, drop_duplicates_columns)
                    if sheet_name in target_sheets:
                        existing_data[sheet_name] = pd.concat([existing_data[sheet_name], df]).drop_duplicates(
                            subset=drop_duplicates_columns, keep='first')
                        print(f"已处理sheet: {sheet_name}")
                    else:
                        existing_data[sheet_name] = df
                        if sheet_name not in target_sheets:
                            target_sheets.append(sheet_name)
                        print(f"已处理sheet: {sheet_name}")

                # 写入目标文件中其他未被覆盖的sheet
                for sheet in target_sheets:
                    existing_data[sheet].to_excel(writer, sheet_name=sheet, index=False)
                    print(f"保留原有sheet: {sheet}")
            self.format_excel(main_file_name)
            print("数据合并完成!")

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            print(f"处理出错: {str(e)}")

    def format_excel(self, main_file_name):
        from openpyxl.utils import get_column_letter
        target_sheets = self.get_excel_sheet_names(main_file_name)
        existing_data = {}
        for sheet_name in target_sheets:
            existing_data[sheet_name] = pd.read_excel(main_file_name, sheet_name=sheet_name)

        # 创建一个ExcelWriter对象，重新格式化后写入文件
        with pd.ExcelWriter(main_file_name, engine='openpyxl') as writer:
            # 写入目标文件中其他未被覆盖的sheet并格式化
            for sheet_name, df in existing_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"格式化原有sheet: {sheet_name}")

                # 获取工作表对象
                worksheet = writer.sheets[sheet_name]

                # 自动调整列宽
                for column in df:
                    max_len = min(df[column].astype(str).apply(len).max(), 70)
                    # print(column, max_len)
                    col_letter = get_column_letter(df.columns.get_loc(column) + 1)
                    worksheet.column_dimensions[col_letter].width = max_len + 4


def save_and_append_xlsx(result_df, sheet_name, overwrite_col=None, output_path="./a.xlsx"):
    result_df_s = result_df.copy()
    if "index" not in result_df_s.columns:
        result_df_s.reset_index(inplace=True, drop=True)
    if not os.path.exists(output_path):
        print(f"{output_path}不存在，直接to_excel写入")
        result_df_s.to_excel(output_path, sheet_name=sheet_name, index=False)
    else:
        try:
            book = load_workbook(output_path)
        except:
            print(f"{output_path}文件损坏，直接to_excel写入")
            result_df_s.to_excel(output_path, sheet_name=sheet_name, index=False)
            return
        pandas_v = pd.__version__
        # pandas<1.3.0时if_sheet_exists="replace"不生效
        if pandas_v < "1.3.0":
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                try:
                    if overwrite_col:
                        try:
                            df_e = pd.read_excel(output_path, sheet_name=sheet_name)
                        except:
                            df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
                        df_e = df_e[~df_e[overwrite_col].isin(set(result_df_s[overwrite_col].tolist()))]
                        result_df_s = pd.concat([result_df_s, df_e])
                except:
                    pass
                if sheet_name in writer.sheets:
                    # 清空现有工作表
                    worksheet = writer.sheets[sheet_name]
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                                   max_col=worksheet.max_column):
                        for cell in row:
                            cell.value = None
                result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)
            # 保存更改
            writer.save()

        else:
            with pd.ExcelWriter(output_path, engine='openpyxl', mode="a", if_sheet_exists="replace") as writer:
                try:
                    if overwrite_col:
                        try:
                            df_e = pd.read_excel(output_path, sheet_name=sheet_name)
                        except:
                            df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
                        df_e = df_e[~df_e[overwrite_col].isin(set(result_df_s[overwrite_col].tolist()))]
                        print(f"{output_path} 过滤历史数据，然后写入")
                        result_df_s = pd.concat([result_df_s, df_e])
                    result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as e:
                    print("save_and_append_xlsx exception:", e)
                    result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)


if __name__ == '__main__':
    mode = "test"
    engine = ExcelEngine()
    sub_file_name, main_file_name = engine.get_summary_path("bili", "2025-01-05")
    drop_duplicates_column = '标题'

    engine.merge_excel(sub_file_name, main_file_name, drop_duplicates_column)
